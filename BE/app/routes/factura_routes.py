"""
Rutas de API para operaciones de Facturas.
Proporciona endpoints CRUD y de descarga de facturas.
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
from uuid import UUID

from BE.app.db import get_db
from BE.app.models.factura_models import Factura
from BE.app.models.transaccion import Transaccion
from BE.app.schemas.factura_schemas import (
    FacturaCreate, 
    FacturaUpdate, 
    FacturaOut, 
    FacturaResumen,
    FacturaConDetallesCreate
)
from BE.app.services.facturacion_service import (
    crear_factura,
    obtener_factura_por_id,
    listar_facturas,
    actualizar_factura,
    eliminar_factura,
    obtener_estadisticas_facturacion,
    obtener_top_clientes
)

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io

router = APIRouter(prefix="/api/facturas", tags=["Facturas"])


# =========================================================
# 🟦 CREAR FACTURA SIMPLE (LEGACY)
# =========================================================
@router.post("/", response_model=FacturaOut, status_code=status.HTTP_201_CREATED)
def crear_nueva_factura(
    factura: FacturaCreate,
    db: Session = Depends(get_db)
):
    """
    Crea una nueva factura simple (legacy).
    Soporta cliente manual o id_cliente de tabla normalizada.
    """
    return crear_factura(db, factura)


# =========================================================
# 🆕 CREAR FACTURA CON DETALLES (NORMALIZADA)
# =========================================================
@router.post("/con-detalles", response_model=FacturaOut, status_code=status.HTTP_201_CREATED)
def crear_factura_con_detalles(
    factura_data: FacturaConDetallesCreate,
    db: Session = Depends(get_db)
):
    """
    Crea factura normalizada con múltiples líneas de productos/servicios.
    - Valida cliente y productos
    - Calcula totales automáticamente
    - Actualiza inventario de productos físicos
    - Crea líneas en tabla factura_detalle
    """
    # Convertir a FacturaCreate
    factura_create = FacturaCreate(
        id_cliente=factura_data.id_cliente,
        condiciones_pago=factura_data.condiciones_pago,
        vendedor=factura_data.vendedor,
        fecha_vencimiento=factura_data.fecha_vencimiento,
        notas=factura_data.notas,
        descuento=factura_data.descuento_global
    )
    
    # Convertir detalles a formato dict
    detalles = [detalle.dict() for detalle in factura_data.detalles]
    
    return crear_factura(db, factura_create, detalles)


# =========================================================
# 🟦 LISTAR FACTURAS
# =========================================================
@router.get("/", response_model=List[FacturaResumen])
def listar_todas_facturas(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    cliente: Optional[str] = Query(None),
    fecha_desde: Optional[datetime] = Query(None),
    fecha_hasta: Optional[datetime] = Query(None),
    db: Session = Depends(get_db)
):
    """Lista facturas con filtros opcionales"""
    facturas = listar_facturas(
        db, skip, limit, cliente, fecha_desde, fecha_hasta
    )
    
    return [
        FacturaResumen(
            id_factura=f.id_factura,
            numero_factura=f.numero_factura,
            cliente=f.cliente_obj.nombre if f.cliente_obj else (f.cliente or "Sin cliente"),
            monto_total=f.monto_total,
            fecha_emision=f.fecha_emision
        )
        for f in facturas
    ]


# =========================================================
# 🟦 OBTENER FACTURA POR ID
# =========================================================
@router.get("/{factura_id}", response_model=FacturaOut)
def obtener_una_factura(
    factura_id: UUID,
    db: Session = Depends(get_db)
):
    """Obtiene una factura específica por ID"""
    factura = obtener_factura_por_id(db, factura_id)
    
    if not factura:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Factura {factura_id} no encontrada"
        )
    
    return factura


# =========================================================
# 🟦 ACTUALIZAR FACTURA
# =========================================================
@router.put("/{factura_id}", response_model=FacturaOut)
def actualizar_una_factura(
    factura_id: UUID,
    factura_update: FacturaUpdate,
    db: Session = Depends(get_db)
):
    """Actualiza una factura existente"""
    return actualizar_factura(db, factura_id, factura_update)


# =========================================================
# 🟦 ELIMINAR FACTURA
# =========================================================
@router.delete("/{factura_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_una_factura(
    factura_id: UUID,
    db: Session = Depends(get_db)
):
    """Elimina una factura"""
    eliminar_factura(db, factura_id)
    return None


# =========================================================
# 🟦 ESTADÍSTICAS DE FACTURACIÓN
# =========================================================
@router.get("/estadisticas/resumen", response_model=dict)
def estadisticas_facturacion(
    fecha_desde: Optional[datetime] = Query(None),
    fecha_hasta: Optional[datetime] = Query(None),
    db: Session = Depends(get_db)
):
    """Obtiene estadísticas de facturación"""
    return obtener_estadisticas_facturacion(db, fecha_desde, fecha_hasta)


# =========================================================
# 🟦 TOP CLIENTES
# =========================================================
@router.get("/estadisticas/top-clientes", response_model=List[dict])
def top_clientes(
    limite: int = Query(10, ge=1, le=100),
    fecha_desde: Optional[datetime] = Query(None),
    fecha_hasta: Optional[datetime] = Query(None),
    db: Session = Depends(get_db)
):
    """Obtiene los clientes con más compras"""
    return obtener_top_clientes(db, limite, fecha_desde, fecha_hasta)


# =========================================================
# 🟦 DESCARGAR FACTURA EN PDF
# =========================================================
@router.get("/{factura_id}/descargar-pdf")
def descargar_factura_pdf(
    factura_id: UUID,
    db: Session = Depends(get_db)
):
    """Genera y descarga factura en formato PDF profesional"""
    factura = obtener_factura_por_id(db, factura_id)
    
    if not factura:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Factura {factura_id} no encontrada"
        )
    
    # Crear PDF en memoria
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # ========== ENCABEZADO ==========
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=10,
        alignment=1,  # Centrado
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#666666'),
        alignment=1,
        spaceAfter=20
    )
    
    title = Paragraph("FACTURA", title_style)
    elements.append(title)
    elements.append(Paragraph(f"No. {factura.numero_factura}", subtitle_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # ========== INFORMACIÓN DE LA EMPRESA (Placeholder) ==========
    empresa_data = [[
        Paragraph("<b>EMPRESA S.A. DE C.V.</b><br/>NIT: 0000-000000-000-0<br/>Dirección: Ciudad Capital<br/>Tel: (000) 0000-0000", styles['Normal'])
    ]]
    empresa_table = Table(empresa_data, colWidths=[6*inch])
    empresa_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f0f0')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1f4788')),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(empresa_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # ========== INFORMACIÓN DE CLIENTE Y FACTURA ==========
    info_data = [
        [Paragraph("<b>CLIENTE:</b>", styles['Normal']), factura.cliente or "N/A"],
        [Paragraph("<b>NIT:</b>", styles['Normal']), factura.nit_cliente or "C/F"],
        [Paragraph("<b>Dirección:</b>", styles['Normal']), factura.direccion_cliente or "N/A"],
        [Paragraph("<b>Teléfono:</b>", styles['Normal']), factura.telefono_cliente or "N/A"],
        [Paragraph("<b>Email:</b>", styles['Normal']), factura.email_cliente or "N/A"],
        ["", ""],
        [Paragraph("<b>Fecha Emisión:</b>", styles['Normal']), factura.fecha_emision.strftime("%d/%m/%Y %H:%M") if factura.fecha_emision else "N/A"],
        [Paragraph("<b>Fecha Vencimiento:</b>", styles['Normal']), factura.fecha_vencimiento.strftime("%d/%m/%Y") if factura.fecha_vencimiento else "N/A"],
        [Paragraph("<b>Condiciones Pago:</b>", styles['Normal']), factura.condiciones_pago or "Contado"],
        [Paragraph("<b>Vendedor:</b>", styles['Normal']), factura.vendedor or "N/A"],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8eef7')),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('LINEABOVE', (0, 6), (-1, 6), 2, colors.HexColor('#1f4788')),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # ========== TOTALES ==========
    totales_header_style = ParagraphStyle('TotalesHeader', parent=styles['Normal'], fontSize=11, fontName='Helvetica-Bold', textColor=colors.whitesmoke)
    totales_data = [
        [Paragraph("CONCEPTO", totales_header_style), Paragraph("MONTO", totales_header_style)],
        ["Subtotal", f"${float(factura.subtotal):,.2f}"],
        ["Descuento", f"-${float(factura.descuento):,.2f}"],
        ["IVA (13%)", f"${float(factura.iva):,.2f}"],
        ["", ""],
        [Paragraph("<b>TOTAL A PAGAR</b>", styles['Normal']), Paragraph(f"<b>${float(factura.monto_total):,.2f}</b>", styles['Normal'])],
    ]
    
    totales_table = Table(totales_data, colWidths=[4*inch, 2*inch])
    totales_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8eef7')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
    ]))
    
    elements.append(totales_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # ========== NOTAS ==========
    if factura.notas:
        elements.append(Paragraph("<b>NOTAS:</b>", styles['Normal']))
        elements.append(Spacer(1, 0.1*inch))
        notas_style = ParagraphStyle('Notas', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#666666'))
        elements.append(Paragraph(factura.notas, notas_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # ========== PIE DE PÁGINA ==========
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#999999'), alignment=1)
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph("____________________________________", footer_style))
    elements.append(Paragraph("Firma y Sello", footer_style))
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph(f"Documento generado electrónicamente - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", footer_style))
    
    # Generar PDF
    doc.build(elements)
    pdf_buffer.seek(0)
    
    return StreamingResponse(
        iter([pdf_buffer.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Factura_{factura.numero_factura}.pdf"}
    )


# =========================================================
# 🟦 DESCARGAR FACTURA EN EXCEL
# =========================================================
@router.get("/{factura_id}/descargar-excel")
def descargar_factura_excel(
    factura_id: UUID,
    db: Session = Depends(get_db)
):
    """Genera y descarga factura en formato Excel completo"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl no está instalado"
        )
    
    factura = obtener_factura_por_id(db, factura_id)
    
    if not factura:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Factura {factura_id} no encontrada"
        )
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factura"
    
    # Estilos
    title_font = Font(size=18, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
    header_font = Font(size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    label_font = Font(size=10, bold=True)
    total_font = Font(size=12, bold=True)
    total_fill = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== TÍTULO ==========
    ws.merge_cells('A1:F1')
    ws['A1'] = "FACTURA"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Número de factura
    ws.merge_cells('A2:F2')
    ws['A2'] = f"No. {factura.numero_factura}"
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # ========== DATOS DE EMPRESA (Placeholder) ==========
    current_row = 4
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "EMPRESA S.A. DE C.V. - NIT: 0000-000000-000-0"
    ws[f'A{current_row}'].font = Font(size=11, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    ws[f'A{current_row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    # ========== INFORMACIÓN DE CLIENTE ==========
    current_row = 6
    ws[f'A{current_row}'] = "INFORMACIÓN DEL CLIENTE"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:F{current_row}')
    
    cliente_data = [
        ("Cliente:", factura.cliente or "N/A"),
        ("NIT:", factura.nit_cliente or "C/F"),
        ("Dirección:", factura.direccion_cliente or "N/A"),
        ("Teléfono:", factura.telefono_cliente or "N/A"),
        ("Email:", factura.email_cliente or "N/A"),
    ]
    
    current_row += 1
    for label, value in cliente_data:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = label_font
        ws[f'A{current_row}'].border = thin_border
        ws[f'B{current_row}'] = value
        ws[f'B{current_row}'].border = thin_border
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
    
    # ========== INFORMACIÓN DE FACTURA ==========
    current_row += 1
    ws[f'A{current_row}'] = "DETALLES DE LA FACTURA"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:F{current_row}')
    
    factura_data = [
        ("Fecha Emisión:", factura.fecha_emision.strftime("%d/%m/%Y %H:%M") if factura.fecha_emision else "N/A"),
        ("Fecha Vencimiento:", factura.fecha_vencimiento.strftime("%d/%m/%Y") if factura.fecha_vencimiento else "N/A"),
        ("Condiciones Pago:", factura.condiciones_pago or "Contado"),
        ("Vendedor:", factura.vendedor or "N/A"),
    ]
    
    current_row += 1
    for label, value in factura_data:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = label_font
        ws[f'A{current_row}'].border = thin_border
        ws[f'B{current_row}'] = value
        ws[f'B{current_row}'].border = thin_border
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
    
    # ========== TOTALES ==========
    current_row += 2
    ws[f'D{current_row}'] = "CONCEPTO"
    ws[f'E{current_row}'] = "MONTO"
    ws[f'D{current_row}'].font = header_font
    ws[f'E{current_row}'].font = header_font
    ws[f'D{current_row}'].fill = header_fill
    ws[f'E{current_row}'].fill = header_fill
    ws[f'D{current_row}'].alignment = Alignment(horizontal='right')
    ws[f'E{current_row}'].alignment = Alignment(horizontal='right')
    
    totales = [
        ("Subtotal", float(factura.subtotal)),
        ("Descuento", -float(factura.descuento)),
        ("IVA (13%)", float(factura.iva)),
        ("TOTAL A PAGAR", float(factura.monto_total)),
    ]
    
    current_row += 1
    for label, monto in totales:
        ws[f'D{current_row}'] = label
        ws[f'E{current_row}'] = monto
        ws[f'E{current_row}'].number_format = '$#,##0.00'
        ws[f'D{current_row}'].border = thin_border
        ws[f'E{current_row}'].border = thin_border
        ws[f'D{current_row}'].alignment = Alignment(horizontal='right')
        ws[f'E{current_row}'].alignment = Alignment(horizontal='right')
        
        if label == "TOTAL A PAGAR":
            ws[f'D{current_row}'].font = total_font
            ws[f'E{current_row}'].font = total_font
            ws[f'D{current_row}'].fill = total_fill
            ws[f'E{current_row}'].fill = total_fill
        
        current_row += 1
    
    # ========== NOTAS ==========
    if factura.notas:
        current_row += 2
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "NOTAS:"
        ws[f'A{current_row}'].font = label_font
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = factura.notas
        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True)
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Guardar en memoria
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return StreamingResponse(
        iter([excel_buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Factura_{factura.numero_factura}.xlsx"}
    )